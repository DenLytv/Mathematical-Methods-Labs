import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
import matplotlib.pyplot as plt
from prettytable import PrettyTable


class Alternative:
    def __init__(self, name: str, q1: int, q2: int, q3: int = None):
        self.name = name
        self.q1 = q1
        self.q2 = q2
        self.q3 = q3
        self.pareto = None
        self.slater = None


def compare_by_pareto(first: Alternative, second: Alternative):
    comparisons = [
        first.q1 >= second.q1,
        first.q2 >= second.q2,
    ]
    if first.q3 is not None and second.q3 is not None:
        comparisons.append(first.q3 >= second.q3)

    strict_better = (
            (first.q1 > second.q1) or
            (first.q2 > second.q2) or
            (first.q3 is not None and second.q3 is not None and first.q3 > second.q3)
    )
    return all(comparisons) and strict_better


def compare_by_slater(first: Alternative, second: Alternative):
    comparisons = [
        first.q1 > second.q1,
        first.q2 > second.q2,
    ]
    if first.q3 is not None and second.q3 is not None:
        comparisons.append(first.q3 > second.q3)

    return all(comparisons)


if __name__ == '__main__':
    print("Select input mode:")
    print("1 - Two-digit numbers (e.g., 32, 58, ...)")
    print("2 - Three-digit numbers (e.g., 463, 478, ...)")
    mode = input("> ").strip()

    nums_string = input("Enter numbers separated by a space: ")
    nums_list = nums_string.split()

    a_list = []
    for i, num in enumerate(nums_list):
        if mode == '1':
            if len(num) != 2:
                print(f"Invalid input '{num}' for two-digit mode. Expected two digits.")
                exit(1)
            q1 = int(num[0])
            q2 = int(num[1])
            a_list.append(Alternative(name=f"A{i + 1}", q1=q1, q2=q2))
        elif mode == '2':
            if len(num) != 3:
                print(f"Invalid input '{num}' for three-digit mode. Expected three digits.")
                exit(1)
            q1 = int(num[0])
            q2 = int(num[1])
            q3 = int(num[2])
            a_list.append(Alternative(name=f"A{i + 1}", q1=q1, q2=q2, q3=q3))
        else:
            print("Invalid mode selected.")
            exit(1)

    a_count = len(a_list)
    for i in range(a_count):
        for j in range(a_count):
            if i == j:
                continue

            if not a_list[j].pareto and compare_by_pareto(a_list[i], a_list[j]):
                a_list[j].pareto = f"A{i + 1}"

            if not a_list[j].slater and compare_by_slater(a_list[i], a_list[j]):
                a_list[j].slater = f"A{i + 1}"

    pareto_groups = {}
    slater_groups = {}

    for alt in a_list:
        key = (alt.q1, alt.q2) if alt.q3 is None else (alt.q1, alt.q2, alt.q3)
        if not alt.pareto:
            if key not in pareto_groups:
                pareto_groups[key] = []
            pareto_groups[key].append(alt.name)
        if not alt.slater:
            if key not in slater_groups:
                slater_groups[key] = []
            slater_groups[key].append(alt.name)

    pareto_res = ["=".join(group) for group in pareto_groups.values()]
    slater_res = ["=".join(group) for group in slater_groups.values()]

    table = PrettyTable()
    table.title = "Results"

    if mode == '1':
        table.field_names = ["Alternative", "Q1", "Q2", "Pareto", "Slater"]
        for alt in a_list:
            table.add_row([alt.name, alt.q1, alt.q2, alt.pareto or '-', alt.slater or '-'])
    elif mode == '2':
        table.field_names = ["Alternative", "Q1", "Q2", "Q3", "Pareto", "Slater"]
        for alt in a_list:
            table.add_row([alt.name, alt.q1, alt.q2, alt.q3, alt.pareto or '-', alt.slater or '-'])

    print(table)

    print(f"Pareto: " + ", ".join(pareto_res))
    print(f"Slater: " + ", ".join(slater_res))

    while True:
        print("\nSave data to a table? (Y/N)")
        answer = input("> ").lower().strip()

        if answer == "y":

            FILE_PATH = "multi_criteria_results.xlsx"

            try:
                if mode == '1':
                    data = {'Q/A': ['Q1', 'Q2', 'Pareto', 'Slater']}
                    for alt in a_list:
                        data[alt.name] = [alt.q1, alt.q2, alt.pareto or '-', alt.slater or '-']
                elif mode == '2':
                    data = {'Q/A': ['Q1', 'Q2', 'Q3', 'Pareto', 'Slater']}
                    for alt in a_list:
                        data[alt.name] = [alt.q1, alt.q2, alt.q3, alt.pareto or '-', alt.slater or '-']

                df = pd.DataFrame(data)
                df.to_excel(FILE_PATH, index=False)

                wb = load_workbook(FILE_PATH)
                ws = wb.active

                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))

                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                wb.save(FILE_PATH)
                print(f"File {FILE_PATH} saved")
            except PermissionError:
                print(f"ERROR: File {FILE_PATH} is not accessible!\n\tThis file may be open now!")
            break
        elif answer == "n":
            break

        print("\nWrong input! Try again.")

    while True:
        print("\nDraw graphs? (Y/N)")
        answer = input("> ").lower().strip()

        if answer == "y":
            def setup_subplot(ax, connected_groups, color='r-', title=''):
                point_coords = {}

                for alt in a_list:
                    coords = (alt.q1, alt.q2) if alt.q3 is None else (alt.q1, alt.q2)
                    if coords not in point_coords:
                        point_coords[coords] = []
                    point_coords[coords].append(alt.name)

                for coords, alts in point_coords.items():
                    x, y = coords
                    ax.scatter(x, y, color=color[0], zorder=3)
                    ax.text(x + 0.2, y + 0.2, "=".join(alts), fontsize=9, zorder=3)

                if len(connected_groups) > 1:
                    sorted_groups = sorted(connected_groups, key=lambda p: (p[0], -p[1]))
                    x_vals, y_vals = zip(*[(p[0], p[1]) for p in sorted_groups])

                    ax.plot(x_vals, y_vals, color, linewidth=2, zorder=2)

                ax.set_title(title)
                ax.grid(True, linestyle='--', linewidth=0.5, alpha=0.7)
                ax.set_xlabel("Q1")
                ax.set_ylabel("Q2")


            fig, axs = plt.subplots(1, 2, figsize=(15, 7))
            setup_subplot(axs[0], list(pareto_groups.keys()), 'r-', 'Pareto limit')
            setup_subplot(axs[1], list(slater_groups.keys()), 'b-', 'Slater limit')
            plt.tight_layout()
            plt.show()

            break
        elif answer == "n":
            break

        print("\nWrong input! Try again.")

        # 32 82 21 66 63 60 82 11 85 86 85 30 90 83 14
        # 28 39 25 90 36 60 18 43 37 28 82 21 10 55 88
        # 08 22 83 50 57 97 27 26 69 71 51 49 10 28 39